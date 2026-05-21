import os
import json
from openpyxl import load_workbook, Workbook

# Zuordnung Ordner → Tabellenblatt
CATEGORY_MAP = {
    "NP-Betrieb": "NP-Betrieb",
    "Ö-Arbeit": "Ö-Arbeit",
    "Fairteiler": "Fairteiler",
    "Anlieferung": "Anlieferung",
    "geschlossen": "geschlossen"
}

EXCEL_PATH = "excel/foodsharing.xlsx"
DATA_DIR = "data"

def ensure_sheet(wb, name):
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(["Titel JSON-Dokument", "", "Datum", "ID", "Name"])
    return wb[name]

def extract_info(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Titel = Dateiname
    title = os.path.basename(json_path)

    # Datum, ID, Name aus JSON extrahieren
    # (du kannst das später anpassen, wenn du mehr Felder brauchst)
    date = data.get("date", "")
    store_id = data.get("storeId", "")
    name = data.get("name", "")

    return title, date, store_id, name

def main():
    # Excel laden oder neu erstellen
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Alle JSON-Dateien durchgehen
    for root, dirs, files in os.walk(DATA_DIR):
        for file in files:
            if not file.endswith(".json"):
                continue

            json_path = os.path.join(root, file)

            # Kategorie aus Ordnername bestimmen
            category = os.path.basename(os.path.dirname(json_path))
            sheet_name = CATEGORY_MAP.get(category)

            if not sheet_name:
                print(f"Unbekannte Kategorie: {category}")
                continue

            ws = ensure_sheet(wb, sheet_name)

            title, date, store_id, name = extract_info(json_path)

            ws.append([title, "", date, store_id, name])

    wb.save(EXCEL_PATH)
    print("Excel erfolgreich aktualisiert.")

if __name__ == "__main__":
    main()
