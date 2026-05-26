import os
import json
from openpyxl import load_workbook, Workbook

# Zuordnung Ordner → Tabellenblatt
CATEGORY_MAP = {
    "NP-Betrieb": "NP-Betrieb",
    "Ö-Arbeit": "Ö-Arbeit",
    "Fairteiler": "Fairteiler",
    "Anlieferung": "Anlieferung",
    "geschlossen": "geschlossen"
}

EXCEL_PATH = "excel/foodsharing.xlsx"
DATA_DIR = "data"

def ensure_sheet(wb, name):
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(["Titel JSON-Dokument", "", "Datum", "ID", "Name"])
    return wb[name]

def extract_info(json_path):
    filename = os.path.basename(json_path)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Immer eine Liste daraus machen
    if isinstance(data, dict):
        data = [data]

    results = []

    for entry in data:
        date = entry.get("date", "")

        profile = entry.get("profile", {})
        store_id = profile.get("id", "")
        name = profile.get("name", "")

        # Titel = Dateiname
        title = filename

        results.append((title, date, store_id, name))

    return results





def main():
    # Excel laden oder neu erstellen
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Alle JSON-Dateien durchgehen
    for root, dirs, files in os.walk(DATA_DIR):
        for file in files:
            if not file.endswith(".json"):
                continue

            json_path = os.path.join(root, file)

            # Kategorie aus Ordnername bestimmen
            category = os.path.basename(os.path.dirname(json_path))
            sheet_name = CATEGORY_MAP.get(category)

            if not sheet_name:
                print(f"Unbekannte Kategorie: {category}")
                continue

            ws = ensure_sheet(wb, sheet_name)

            # ALLE Einträge aus der JSON-Datei holen
            entries = extract_info(json_path)

            # Jede Aktion einzeln eintragen
            for title, date, store_id, name in entries:
                ws.append([title, "", date, store_id, name])

    wb.save(EXCEL_PATH)
    print("Excel erfolgreich aktualisiert.")


if __name__ == "__main__":
    main()
